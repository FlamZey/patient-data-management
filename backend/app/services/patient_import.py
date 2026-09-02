"""Parses and validates an uploaded patient spreadsheet before anything is
encrypted or written to the DB. Pure Python in and out (raw bytes + filename
in, a plain dataclass result out) -- no FastAPI or DB imports, so this is
independently unit-testable and reusable regardless of how the file arrives.
"""

import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from functools import partial
from io import BytesIO
from typing import Any, Generator, Iterable, Literal, get_args

import openpyxl
import xlrd
from email_validator import EmailNotValidError, validate_email

from app.core.text import strip_invisible

REQUIRED_COLUMNS = ["Patient ID", "First Name", "Last Name", "Date of Birth", "Gender"]

# The Literal is the single source of truth for allowed values -- ALLOWED_GENDERS
# is derived from it for runtime membership checks, and schemas.PatientUpdate
# imports the Literal itself for its own type annotation, so the two never drift.
Gender = Literal["Male", "Female", "Other", "Prefer not to say"]
ALLOWED_GENDERS: tuple[str, ...] = get_args(Gender)

MaritalStatus = Literal["Single", "Married", "Divorced", "Widowed", "Separated", "Domestic Partnership"]
ALLOWED_MARITAL_STATUSES: tuple[str, ...] = get_args(MaritalStatus)

RaceEthnicity = Literal[
    "White",
    "Black or African American",
    "Asian",
    "Hispanic or Latino",
    "American Indian or Alaska Native",
    "Native Hawaiian or Other Pacific Islander",
    "Middle Eastern or North African",
    "Two or More Races",
    "Prefer not to say",
]
ALLOWED_RACE_ETHNICITIES: tuple[str, ...] = get_args(RaceEthnicity)

EmergencyContactRelationship = Literal[
    "Spouse",
    "Parent",
    "Sibling",
    "Child",
    "Friend",
    "Partner",
    "Grandparent",
    "Grandchild",
    "Caregiver",
    "Other Relative",
]
ALLOWED_RELATIONSHIPS: tuple[str, ...] = get_args(EmergencyContactRelationship)

BloodType = Literal["O+", "A+", "B+", "AB+", "O-", "A-", "B-", "AB-"]
ALLOWED_BLOOD_TYPES: tuple[str, ...] = get_args(BloodType)

SmokingStatus = Literal[
    "Never smoker",
    "Former smoker",
    "Current every day smoker",
    "Current some day smoker",
    "Light tobacco smoker",
    "Heavy tobacco smoker",
    "Smoker, current status unknown",
    "Unknown if ever smoked",
]
ALLOWED_SMOKING_STATUSES: tuple[str, ...] = get_args(SmokingStatus)

AlcoholUse = Literal["Never", "Rarely", "Occasional", "Moderate", "Heavy", "In recovery"]
ALLOWED_ALCOHOL_USE: tuple[str, ...] = get_args(AlcoholUse)

CareDepartment = Literal[
    "Primary Care",
    "Pediatrics",
    "Cardiology",
    "Endocrinology",
    "Pulmonology",
    "Orthopedics",
    "Psychiatry",
    "Nephrology",
    "General Medicine",
]
ALLOWED_CARE_DEPARTMENTS: tuple[str, ...] = get_args(CareDepartment)

# Every value Faker's state_abbr() can emit for this repo's pinned Faker version: 50 states + DC
# + US territories/freely-associated states -- confirmed empirically rather than hand-guessed, so
# the sample generator and this validator can never disagree.
ALLOWED_STATE_CODES: tuple[str, ...] = (
    "AK", "AL", "AR", "AS", "AZ", "CA", "CO", "CT", "DC", "DE", "FL", "FM", "GA", "GU", "HI",
    "IA", "ID", "IL", "IN", "KS", "KY", "LA", "MA", "MD", "ME", "MH", "MI", "MN", "MO", "MP",
    "MS", "MT", "NC", "ND", "NE", "NH", "NJ", "NM", "NV", "NY", "OH", "OK", "OR", "PA", "PR",
    "PW", "RI", "SC", "SD", "TN", "TX", "UT", "VA", "VI", "VT", "WA", "WI", "WV", "WY",
)

# Optional columns: unlike REQUIRED_COLUMNS, a workbook may include any subset of these (including
# none) and still validate. This list is the single source of truth for both the upload validator
# and scripts/generate_load_test_workbook.py's preview generator, so the two can't drift apart.
OPTIONAL_COLUMNS = [
    "Street Address",
    "City",
    "State",
    "Zip",
    "Phone",
    "Email",
    "Emergency Contact Name",
    "Emergency Contact Relationship",
    "Emergency Contact Phone",
    "Preferred Language",
    "Race/Ethnicity",
    "Marital Status",
    "Occupation",
    "Insurance Provider",
    "Policy Number",
    "PCP Name",
    "Care Department",
    "Registration Date",
    "Last Visit Date",
    "Preferred Pharmacy",
    "Blood Type",
    "Height (in)",
    "Weight (lbs)",
    "Systolic BP",
    "Diastolic BP",
    "Allergies",
    "Current Medications",
    "Chronic Conditions (ICD-10)",
    "Immunization History",
    "Smoking Status",
    "Alcohol Use",
]

# Excel header label -> snake_case field name used in the accepted-row dict, PatientRead/
# PatientUpdate, and the Patient model's <name>_enc columns.
_OPTIONAL_FIELD_NAMES: dict[str, str] = {
    "Street Address": "street_address",
    "City": "city",
    "State": "state",
    "Zip": "zip_code",
    "Phone": "phone",
    "Email": "email",
    "Emergency Contact Name": "emergency_contact_name",
    "Emergency Contact Relationship": "emergency_contact_relationship",
    "Emergency Contact Phone": "emergency_contact_phone",
    "Preferred Language": "preferred_language",
    "Race/Ethnicity": "race_ethnicity",
    "Marital Status": "marital_status",
    "Occupation": "occupation",
    "Insurance Provider": "insurance_provider",
    "Policy Number": "policy_number",
    "PCP Name": "pcp_name",
    "Care Department": "care_department",
    "Registration Date": "registration_date",
    "Last Visit Date": "last_visit_date",
    "Preferred Pharmacy": "preferred_pharmacy",
    "Blood Type": "blood_type",
    "Height (in)": "height_in",
    "Weight (lbs)": "weight_lbs",
    "Systolic BP": "systolic_bp",
    "Diastolic BP": "diastolic_bp",
    "Allergies": "allergies",
    "Current Medications": "current_medications",
    "Chronic Conditions (ICD-10)": "chronic_conditions",
    "Immunization History": "immunization_history",
    "Smoking Status": "smoking_status",
    "Alcohol Use": "alcohol_use",
}

# The snake_case field names, in the same order as OPTIONAL_COLUMNS -- shared with
# app.routers.patients so it can build its field->column map without re-listing every name.
OPTIONAL_FIELD_NAMES: tuple[str, ...] = tuple(_OPTIONAL_FIELD_NAMES[label] for label in OPTIONAL_COLUMNS)

# Field-kind classification shared with app.routers.patients, which needs it to know how to
# serialize a value to a string before encryption (and parse it back after decryption).
INT_FIELDS = {"height_in", "weight_lbs", "systolic_bp", "diastolic_bp"}
MULTI_VALUE_FIELDS = {"allergies", "current_medications", "chronic_conditions", "immunization_history"}

# How often _stream_parsed_rows yields a validation progress tuple. Every
# yield is a real chunk sent over the HTTP response it eventually feeds
# (see routers/patients.py) -- yielding every single row measurably slowed
# a 10,000-row upload down (~11s baseline -> ~14s), so this batches progress
# reporting the same way the router already batches its own write-phase
# progress, instead of reporting every row.
VALIDATION_PROGRESS_INTERVAL = 100

# Data rows per upload, not counting the header. The 10MB request-size cap
# (routers/patients.py) doesn't bound this at all: .xlsx is compressed XML,
# and a plain patient roster compresses roughly 15x, so a file well under
# that cap can still carry hundreds of thousands of rows. Enforced while
# reading (below), not after -- the read itself, before a single row is
# validated or encrypted, is what dominates the cost of an oversized file.
MAX_UPLOAD_ROWS = 50_000

MAX_AGE_YEARS = 130
MIN_REGISTRATION_DATE = date(1900, 1, 1)
MIN_HEIGHT_IN, MAX_HEIGHT_IN = 12, 108
MIN_WEIGHT_LBS, MAX_WEIGHT_LBS = 1, 700
# Wide enough to admit a hypertensive-crisis or hypotensive reading without
# rejecting it, but still a real physiologic bound -- catches fat-fingered
# entries (e.g. a 3-digit diastolic) rather than modeling clinical norms.
MIN_SYSTOLIC_BP, MAX_SYSTOLIC_BP = 60, 250
MIN_DIASTOLIC_BP, MAX_DIASTOLIC_BP = 30, 150

_PATIENT_ID_PATTERN = re.compile(r"^[A-Za-z0-9-]+$")
_ZIP_PATTERN = re.compile(r"^\d{5}(-\d{4})?$")
_PHONE_PATTERN = re.compile(r"^[+]?[0-9()\-.\sx]{7,25}$", re.IGNORECASE)
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")
_DATE_STRING_FORMATS = ("%Y-%m-%d", "%m/%d/%Y")


class PatientImportError(Exception):
    """Raised when the whole file must be rejected before any row is
    processed: an unsupported extension, an unreadable workbook, or a header
    row that doesn't have exactly the required columns."""


@dataclass
class RejectedRow:
    row: int
    field: str
    reason: str


@dataclass
class PatientImportResult:
    total_rows: int
    accepted: list[dict] = field(default_factory=list)
    rejected: list[RejectedRow] = field(default_factory=list)


def parse_patient_upload(
    *, filename: str, content: bytes, existing_patient_codes: Iterable[str] = ()
) -> PatientImportResult:
    """existing_patient_codes are the Patient.patient_code values already
    owned by the uploading manager -- passed in rather than queried here so
    this module stays DB-free. Raises PatientImportError for whole-file
    failures; per-row failures are collected in the returned result instead.

    A thin, non-streaming wrapper around _stream_parsed_rows, for callers
    (and all of this module's existing tests) that just want the final
    result and don't care about per-row progress -- see
    parse_patient_upload_streaming for the progress-reporting version."""
    column_index, indexed_rows = _read_and_validate_header(filename, content)
    *_, result = _stream_parsed_rows(column_index, indexed_rows, existing_patient_codes)
    return result


def parse_patient_upload_streaming(
    *, filename: str, content: bytes, existing_patient_codes: Iterable[str] = ()
) -> Generator[tuple[int, int] | PatientImportResult, None, None]:
    """Same validation as parse_patient_upload, but yields (processed, total)
    progress tuples as it works through the rows, with the final
    PatientImportResult as its last yielded item (isinstance-check to tell
    the two apart) -- lets a caller (the upload endpoint) report progress on
    a long file instead of blocking silently until everything is done.

    Header/extension validation still happens eagerly, before the first
    yield: a generator's body doesn't run at all until first iterated, but
    that first iteration runs _read_and_validate_header synchronously before
    reaching any `yield`, so a whole-file failure still raises
    PatientImportError on the caller's first `next()`/iteration step, not
    silently swallowed or deferred."""
    column_index, indexed_rows = _read_and_validate_header(filename, content)
    yield from _stream_parsed_rows(column_index, indexed_rows, existing_patient_codes)


def _read_and_validate_header(filename: str, content: bytes) -> tuple[dict[str, int], list[tuple[int, list]]]:
    rows = _read_workbook(filename, content)
    if not rows:
        raise PatientImportError("The file is empty.")

    header_row, *data_rows = rows
    column_index = _validate_header(header_row)

    indexed_rows = [
        (row_number, row)
        for row_number, row in enumerate(data_rows, start=2)
        if _row_has_any_value(row)
    ]
    return column_index, indexed_rows


def _stream_parsed_rows(
    column_index: dict[str, int],
    indexed_rows: list[tuple[int, list]],
    existing_patient_codes: Iterable[str],
) -> Generator[tuple[int, int] | PatientImportResult, None, None]:
    def cell(row: list, name: str) -> Any:
        idx = column_index.get(name)
        if idx is None:
            return None
        return row[idx] if idx < len(row) else None

    code_counts: Counter[str] = Counter()
    for _, row in indexed_rows:
        cleaned = _clean_str(cell(row, "Patient ID"))
        if cleaned:
            code_counts[cleaned] += 1

    existing_codes = set(existing_patient_codes)
    accepted: list[dict] = []
    rejected: list[RejectedRow] = []
    today = date.today()
    total = len(indexed_rows)

    for processed, (row_number, row) in enumerate(indexed_rows, start=1):
        errors: dict[str, str] = {}

        patient_id, error = _validate_patient_id(
            cell(row, "Patient ID"), code_counts=code_counts, existing_codes=existing_codes
        )
        if error:
            errors["Patient ID"] = error

        first_name, error = _validate_name(cell(row, "First Name"), "First Name")
        if error:
            errors["First Name"] = error

        last_name, error = _validate_name(cell(row, "Last Name"), "Last Name")
        if error:
            errors["Last Name"] = error

        date_of_birth, error = _validate_date_of_birth(cell(row, "Date of Birth"), today=today)
        if error:
            errors["Date of Birth"] = error

        gender, error = _validate_gender(cell(row, "Gender"))
        if error:
            errors["Gender"] = error

        optional_values, optional_errors = _validate_optional_fields(row, cell, today, date_of_birth)
        errors.update(optional_errors)

        if errors:
            rejected.extend(
                RejectedRow(row=row_number, field=field_name, reason=reason)
                for field_name, reason in errors.items()
            )
        else:
            existing_codes.add(patient_id)
            accepted.append(
                {
                    "patient_code": patient_id,
                    "first_name": first_name,
                    "last_name": last_name,
                    "date_of_birth": date_of_birth,
                    "gender": gender,
                    **optional_values,
                }
            )

        if processed % VALIDATION_PROGRESS_INTERVAL == 0 or processed == total:
            yield (processed, total)

    yield PatientImportResult(total_rows=total, accepted=accepted, rejected=rejected)


def _read_workbook(filename: str, content: bytes) -> list[list]:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension == "xlsx":
        return _read_xlsx(content)
    if extension == "xls":
        return _read_xls(content)
    raise PatientImportError(f"Unsupported file type for '{filename}'. Only .xlsx and .xls files are accepted.")


def _too_many_rows_error() -> PatientImportError:
    return PatientImportError(
        f"File exceeds the {MAX_UPLOAD_ROWS:,}-row limit. Split it into smaller files and upload each separately."
    )


def _read_xlsx(content: bytes) -> list[list]:
    # data_only=False (the default) is deliberate: a formula cell then reads
    # back as its raw "=..." text instead of Excel's last-cached computed
    # value, so a formula-injection payload can't hide behind a result that
    # looks benign -- our guard needs to see the literal leading character.
    try:
        workbook = openpyxl.load_workbook(BytesIO(content), read_only=True)
    except Exception as exc:
        raise PatientImportError(f"Could not read .xlsx file: {exc}") from exc
    sheet = workbook.worksheets[0]
    # Bails out of the iterator the moment the cap is crossed, rather than
    # materializing the full sheet and rejecting afterward -- read_only=True
    # above only avoids loading the whole workbook into openpyxl's own model;
    # without this check, the list comprehension it used to be still pulled
    # every row into a Python list before anything downstream could object.
    # `rows` includes the header at this point (it's read and counted like
    # any other row, split off by the caller afterward), so the check allows
    # one more stored row than MAX_UPLOAD_ROWS to admit it without counting
    # against the data-row cap.
    rows: list[list] = []
    for row in sheet.iter_rows(values_only=True):
        if len(rows) > MAX_UPLOAD_ROWS:
            raise _too_many_rows_error()
        rows.append(list(row))
    return rows


def _read_xls(content: bytes) -> list[list]:
    try:
        workbook = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise PatientImportError(f"Could not read .xls file: {exc}") from exc
    sheet = workbook.sheet_by_index(0)
    # xlrd has no read_only/streaming mode -- open_workbook above already
    # parsed every row into memory, so the cap here just stops the second,
    # separate per-cell loop below from also running over all of them.
    if sheet.nrows > MAX_UPLOAD_ROWS + 1:  # +1 admits the header row
        raise _too_many_rows_error()
    rows = []
    for r in range(sheet.nrows):
        row = []
        for c in range(sheet.ncols):
            xls_cell = sheet.cell(r, c)
            if xls_cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(xls_cell.value, workbook.datemode))
            elif xls_cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            else:
                row.append(xls_cell.value)
        rows.append(row)
    return rows


def _validate_header(header_row: list) -> dict[str, int]:
    column_index: dict[str, int] = {}
    for idx, raw_name in enumerate(header_row):
        name = _clean_str(raw_name)
        if name:
            column_index[name] = idx

    required = set(REQUIRED_COLUMNS)
    optional = set(OPTIONAL_COLUMNS)
    present = set(column_index)
    missing = required - present
    unexpected = present - required - optional
    if missing or unexpected:
        parts = []
        if missing:
            parts.append(f"missing: {', '.join(sorted(missing))}")
        if unexpected:
            parts.append(f"unexpected: {', '.join(sorted(unexpected))}")
        raise PatientImportError(f"Header row does not match the required columns ({'; '.join(parts)}).")

    return column_index


def _row_has_any_value(row: list) -> bool:
    return any(_clean_str(value) for value in row)


def _clean_str(raw: Any) -> str:
    """Normalises a raw cell/JSON value to a trimmed string.

    Uses strip_invisible rather than str.strip: zero-width characters survive
    strip(), so a value made of nothing but one of them would read as
    non-empty here and sail through every "is this filled in" check below.
    Applied at this single chokepoint so both the upload path and
    PatientUpdate's edit-path validators inherit it.
    """
    if raw is None:
        return ""
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return strip_invisible(str(raw))


def _clean_zip(raw: Any) -> str:
    """Like _clean_str, but restores the leading zero(s) Excel drops when a
    ZIP is entered/formatted as a number (e.g. 02134 -> reads back as 2134.0).
    zfill(5) is a no-op for anything already >=5 characters, including a
    genuinely-too-long value -- that still fails _ZIP_PATTERN as it should."""
    if raw is None:
        return ""
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    if isinstance(raw, int) and not isinstance(raw, bool):
        return str(raw).zfill(5)
    return str(raw).strip()


def _check_formula_injection(raw: Any) -> str | None:
    if isinstance(raw, str) and raw.strip().startswith(_FORMULA_TRIGGER_CHARS):
        return "Value starts with a spreadsheet formula character and is not allowed."
    return None


def _validate_patient_id(
    raw: Any, *, code_counts: Counter, existing_codes: set[str]
) -> tuple[str | None, str | None]:
    value = _clean_str(raw)
    if not value:
        return None, "Patient ID is required."
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    if not _PATIENT_ID_PATTERN.match(value):
        return None, "Patient ID must contain only letters, digits, and hyphens."
    if code_counts[value] > 1:
        return None, "Duplicate Patient ID within this file."
    if value in existing_codes:
        return None, "Patient ID already exists for this manager."
    return value, None


def _validate_name(raw: Any, field_label: str) -> tuple[str | None, str | None]:
    value = _clean_str(raw)
    if not value:
        return None, f"{field_label} is required."
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    return value, None


def _validate_gender(raw: Any) -> tuple[str | None, str | None]:
    value = _clean_str(raw)
    if not value:
        return None, "Gender is required."
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    if value not in ALLOWED_GENDERS:
        return None, f"Gender must be one of: {', '.join(ALLOWED_GENDERS)}."
    return value, None


def _validate_date_of_birth(raw: Any, *, today: date) -> tuple[str | None, str | None]:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, "Date of Birth is required."
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    parsed = _parse_date(raw)
    if parsed is None:
        return None, "Date of Birth must be a valid date (YYYY-MM-DD or MM/DD/YYYY)."
    if parsed > today:
        return None, "Date of Birth cannot be in the future."
    if parsed < _years_before(today, MAX_AGE_YEARS):
        return None, f"Date of Birth cannot be more than {MAX_AGE_YEARS} years in the past."
    return parsed.isoformat(), None


# --- optional-field validators ----------------------------------------------
# Every validator here treats a blank/absent cell as (None, None) -- no error --
# since all of OPTIONAL_COLUMNS is opt-in. Each mirrors the shape of the
# required-field validators above (clean -> blank check -> formula-injection
# check -> domain check) so upload behavior stays consistent across both sets.


def _validate_optional_text(raw: Any, field_label: str) -> tuple[str | None, str | None]:
    value = _clean_str(raw)
    if not value:
        return None, None
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    return value, None


def _validate_enum(raw: Any, field_label: str, allowed: tuple[str, ...]) -> tuple[str | None, str | None]:
    value = _clean_str(raw)
    if not value:
        return None, None
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    if value not in allowed:
        return None, f"{field_label} must be one of: {', '.join(allowed)}."
    return value, None


def _validate_state(raw: Any) -> tuple[str | None, str | None]:
    value = _clean_str(raw).upper()
    if not value:
        return None, None
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    if value not in ALLOWED_STATE_CODES:
        return None, "State must be a valid two-letter USPS state/territory code."
    return value, None


def _validate_zip(raw: Any) -> tuple[str | None, str | None]:
    value = _clean_zip(raw)
    if not value:
        return None, None
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    if not _ZIP_PATTERN.match(value):
        return None, "Zip must be 5 digits, optionally followed by -4 digits."
    return value, None


def _validate_phone(raw: Any, field_label: str) -> tuple[str | None, str | None]:
    value = _clean_str(raw)
    if not value:
        return None, None
    # No separate _check_formula_injection call here (unlike the other validators):
    # a leading "+" is a routine, legitimate first character for a phone number
    # (international dialing prefix), so applying that generic guard would reject
    # real phone data. _PHONE_PATTERN is itself a strict character allowlist --
    # digits, parens, dots, dashes, spaces, and "x" only, no letters besides "x"
    # and no "=", "@", or "|" -- which already rules out any actual formula/
    # function-call injection syntax, making the separate check redundant here.
    if not _PHONE_PATTERN.match(value):
        return None, f"{field_label} must be a valid phone number."
    return value, None


def _validate_email(raw: Any) -> tuple[str | None, str | None]:
    value = _clean_str(raw)
    if not value:
        return None, None
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    try:
        # check_deliverability=False is deliberate: this module does no I/O
        # anywhere else (see the module docstring), and the default would
        # perform a live DNS MX lookup on every row of every upload.
        result = validate_email(value, check_deliverability=False)
    except EmailNotValidError as exc:
        return None, str(exc)
    return result.normalized, None


def _validate_policy_number(raw: Any) -> tuple[str | None, str | None]:
    value = _clean_str(raw)
    if not value:
        return None, None
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    if not _PATIENT_ID_PATTERN.match(value):
        return None, "Policy Number must contain only letters, digits, and hyphens."
    return value, None


def _validate_optional_date(
    raw: Any, field_label: str, *, today: date, min_date: date
) -> tuple[str | None, str | None]:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, None
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    parsed = _parse_date(raw)
    if parsed is None:
        return None, f"{field_label} must be a valid date (YYYY-MM-DD or MM/DD/YYYY)."
    if parsed > today:
        return None, f"{field_label} cannot be in the future."
    if parsed < min_date:
        return None, f"{field_label} cannot be before {min_date.isoformat()}."
    return parsed.isoformat(), None


_validate_registration_date = partial(
    _validate_optional_date, field_label="Registration Date", min_date=MIN_REGISTRATION_DATE
)
_validate_last_visit_date = partial(
    _validate_optional_date, field_label="Last Visit Date", min_date=MIN_REGISTRATION_DATE
)


def _validate_int_range(
    raw: Any, field_label: str, *, min_value: int, max_value: int
) -> tuple[int | None, str | None]:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, None
    formula_error = _check_formula_injection(raw)
    if formula_error:
        return None, formula_error
    # bool is an int subclass in Python -- an Excel TRUE/FALSE cell must not
    # silently become 1/0, so this check has to come before isinstance(raw, int).
    if isinstance(raw, bool):
        return None, f"{field_label} must be a number."
    if isinstance(raw, int):
        value = raw
    elif isinstance(raw, float) and raw.is_integer():
        value = int(raw)
    elif isinstance(raw, str):
        try:
            value = int(raw.strip())
        except ValueError:
            return None, f"{field_label} must be a whole number."
    else:
        return None, f"{field_label} must be a number."
    if value < min_value or value > max_value:
        return None, f"{field_label} must be between {min_value} and {max_value}."
    return value, None


def _split_multi_value_cell(raw: Any) -> list[str]:
    if raw is None:
        return []
    return str(raw).split(",")


def _validate_multi_value_items(items: list[str], field_label: str) -> tuple[list[str] | None, str | None]:
    cleaned: list[str] = []
    for item in items:
        formula_error = _check_formula_injection(item)
        if formula_error:
            return None, formula_error
        value = item.strip()
        if value:
            cleaned.append(value)
    return (cleaned or None), None


def _validate_optional_fields(
    row: list, cell, today: date, date_of_birth: str | None
) -> tuple[dict[str, Any], dict[str, str]]:
    """Validates every OPTIONAL_COLUMNS cell for one row. Returns (values, errors)
    where values is keyed by snake_case field name (always all of OPTIONAL_FIELD_NAMES
    once errors is empty) and errors is keyed by the Excel column label, matching
    the shape RejectedRow expects.

    date_of_birth is the row's already-validated Date of Birth (ISO string), or None
    if that field itself failed validation -- passed in so Registration Date/Last
    Visit Date can be cross-checked against it (see _validate_date_not_before)."""
    values: dict[str, Any] = {}
    errors: dict[str, str] = {}

    def set_field(column_label: str, result: tuple[Any, str | None]) -> None:
        value, error = result
        if error:
            errors[column_label] = error
        else:
            values[_OPTIONAL_FIELD_NAMES[column_label]] = value

    def set_date_field_not_before(
        column_label: str, result: tuple[str | None, str | None], *, floor: str | None, floor_label: str
    ) -> str | None:
        """Like set_field, but for a date field that must also not fall before
        `floor` (another already-validated date) -- a value can pass its own
        format/range validation and still be logically impossible in
        combination with another date (a registration date before the
        patient was born, a last-visit date before the registration date).
        Skipped when `floor` itself is missing/invalid -- nothing valid to
        compare against. Returns the accepted value (or None), so a later
        field can use this one as its own floor."""
        value, error = result
        if error:
            errors[column_label] = error
            return None
        if value is not None and floor is not None and value < floor:
            errors[column_label] = f"{column_label} cannot be before {floor_label}."
            return None
        values[_OPTIONAL_FIELD_NAMES[column_label]] = value
        return value

    set_field("Street Address", _validate_optional_text(cell(row, "Street Address"), "Street Address"))
    set_field("City", _validate_optional_text(cell(row, "City"), "City"))
    set_field("State", _validate_state(cell(row, "State")))
    set_field("Zip", _validate_zip(cell(row, "Zip")))
    set_field("Phone", _validate_phone(cell(row, "Phone"), "Phone"))
    set_field("Email", _validate_email(cell(row, "Email")))
    set_field(
        "Emergency Contact Name",
        _validate_optional_text(cell(row, "Emergency Contact Name"), "Emergency Contact Name"),
    )
    set_field(
        "Emergency Contact Relationship",
        _validate_enum(
            cell(row, "Emergency Contact Relationship"), "Emergency Contact Relationship", ALLOWED_RELATIONSHIPS
        ),
    )
    set_field(
        "Emergency Contact Phone",
        _validate_phone(cell(row, "Emergency Contact Phone"), "Emergency Contact Phone"),
    )
    set_field(
        "Preferred Language", _validate_optional_text(cell(row, "Preferred Language"), "Preferred Language")
    )
    set_field("Race/Ethnicity", _validate_enum(cell(row, "Race/Ethnicity"), "Race/Ethnicity", ALLOWED_RACE_ETHNICITIES))
    set_field("Marital Status", _validate_enum(cell(row, "Marital Status"), "Marital Status", ALLOWED_MARITAL_STATUSES))
    set_field("Occupation", _validate_optional_text(cell(row, "Occupation"), "Occupation"))
    set_field("Insurance Provider", _validate_optional_text(cell(row, "Insurance Provider"), "Insurance Provider"))
    set_field("Policy Number", _validate_policy_number(cell(row, "Policy Number")))
    set_field("PCP Name", _validate_optional_text(cell(row, "PCP Name"), "PCP Name"))
    set_field(
        "Care Department", _validate_enum(cell(row, "Care Department"), "Care Department", ALLOWED_CARE_DEPARTMENTS)
    )
    registration_date = set_date_field_not_before(
        "Registration Date",
        _validate_registration_date(cell(row, "Registration Date"), today=today),
        floor=date_of_birth,
        floor_label="Date of Birth",
    )
    # Last Visit Date must be on/after Registration Date when one was given;
    # falls back to Date of Birth when Registration Date is blank/invalid, so
    # a last-visit date before the patient was ever born is still caught.
    set_date_field_not_before(
        "Last Visit Date",
        _validate_last_visit_date(cell(row, "Last Visit Date"), today=today),
        floor=registration_date if registration_date is not None else date_of_birth,
        floor_label="Registration Date" if registration_date is not None else "Date of Birth",
    )
    set_field("Preferred Pharmacy", _validate_optional_text(cell(row, "Preferred Pharmacy"), "Preferred Pharmacy"))
    set_field("Blood Type", _validate_enum(cell(row, "Blood Type"), "Blood Type", ALLOWED_BLOOD_TYPES))
    set_field(
        "Height (in)",
        _validate_int_range(cell(row, "Height (in)"), "Height (in)", min_value=MIN_HEIGHT_IN, max_value=MAX_HEIGHT_IN),
    )
    set_field(
        "Weight (lbs)",
        _validate_int_range(
            cell(row, "Weight (lbs)"), "Weight (lbs)", min_value=MIN_WEIGHT_LBS, max_value=MAX_WEIGHT_LBS
        ),
    )
    set_field(
        "Systolic BP",
        _validate_int_range(
            cell(row, "Systolic BP"), "Systolic BP", min_value=MIN_SYSTOLIC_BP, max_value=MAX_SYSTOLIC_BP
        ),
    )
    set_field(
        "Diastolic BP",
        _validate_int_range(
            cell(row, "Diastolic BP"), "Diastolic BP", min_value=MIN_DIASTOLIC_BP, max_value=MAX_DIASTOLIC_BP
        ),
    )
    set_field(
        "Allergies",
        _validate_multi_value_items(_split_multi_value_cell(cell(row, "Allergies")), "Allergies"),
    )
    set_field(
        "Current Medications",
        _validate_multi_value_items(_split_multi_value_cell(cell(row, "Current Medications")), "Current Medications"),
    )
    set_field(
        "Chronic Conditions (ICD-10)",
        _validate_multi_value_items(
            _split_multi_value_cell(cell(row, "Chronic Conditions (ICD-10)")), "Chronic Conditions (ICD-10)"
        ),
    )
    set_field(
        "Immunization History",
        _validate_multi_value_items(
            _split_multi_value_cell(cell(row, "Immunization History")), "Immunization History"
        ),
    )
    set_field("Smoking Status", _validate_enum(cell(row, "Smoking Status"), "Smoking Status", ALLOWED_SMOKING_STATUSES))
    set_field("Alcohol Use", _validate_enum(cell(row, "Alcohol Use"), "Alcohol Use", ALLOWED_ALCOHOL_USE))

    return values, errors


def validate_date_of_birth(raw: Any) -> str:
    """Validates a Date of Birth value with the same rules as the upload
    flow (Excel date cell, or ISO/US date string; rejects future dates or
    those more than MAX_AGE_YEARS in the past). Returns the ISO date string
    on success, raises ValueError otherwise -- shared with PatientUpdate in
    schemas.py so manual edits enforce the same rule as bulk upload."""
    value, error = _validate_date_of_birth(raw, today=date.today())
    if error:
        raise ValueError(error)
    return value


def _as_public_validator(private_fn):
    """Wraps a private (value, error) -> tuple validator into a value -> value-or-raise
    function, for use directly as a pydantic @field_validator body. None passes through
    untouched (an unset/cleared field), matching how PatientUpdate treats optional fields."""

    def wrapper(value: Any) -> Any:
        if value is None:
            return None
        result, error = private_fn(value)
        if error:
            raise ValueError(error)
        return result

    return wrapper


validate_first_name = _as_public_validator(partial(_validate_name, field_label="First Name"))
validate_last_name = _as_public_validator(partial(_validate_name, field_label="Last Name"))
validate_street_address = _as_public_validator(partial(_validate_optional_text, field_label="Street Address"))
validate_city = _as_public_validator(partial(_validate_optional_text, field_label="City"))
validate_state = _as_public_validator(_validate_state)
validate_zip_code = _as_public_validator(_validate_zip)
validate_phone = _as_public_validator(partial(_validate_phone, field_label="Phone"))
validate_email_field = _as_public_validator(_validate_email)
validate_emergency_contact_name = _as_public_validator(
    partial(_validate_optional_text, field_label="Emergency Contact Name")
)
validate_emergency_contact_phone = _as_public_validator(
    partial(_validate_phone, field_label="Emergency Contact Phone")
)
validate_preferred_language = _as_public_validator(
    partial(_validate_optional_text, field_label="Preferred Language")
)
validate_occupation = _as_public_validator(partial(_validate_optional_text, field_label="Occupation"))
validate_insurance_provider = _as_public_validator(
    partial(_validate_optional_text, field_label="Insurance Provider")
)
validate_policy_number = _as_public_validator(_validate_policy_number)
validate_pcp_name = _as_public_validator(partial(_validate_optional_text, field_label="PCP Name"))
validate_care_department = _as_public_validator(
    partial(_validate_enum, field_label="Care Department", allowed=ALLOWED_CARE_DEPARTMENTS)
)


def _as_public_date_validator(private_fn):
    """Like _as_public_validator, but for a date validator whose 'today' must be
    evaluated at call time, not bound once at module-import time (server startup)."""

    def wrapper(value: Any) -> Any:
        if value is None:
            return None
        result, error = private_fn(value, today=date.today())
        if error:
            raise ValueError(error)
        return result

    return wrapper


validate_registration_date = _as_public_date_validator(_validate_registration_date)
validate_last_visit_date = _as_public_date_validator(_validate_last_visit_date)

validate_preferred_pharmacy = _as_public_validator(
    partial(_validate_optional_text, field_label="Preferred Pharmacy")
)
validate_height_in = _as_public_validator(
    partial(_validate_int_range, field_label="Height (in)", min_value=MIN_HEIGHT_IN, max_value=MAX_HEIGHT_IN)
)
validate_weight_lbs = _as_public_validator(
    partial(_validate_int_range, field_label="Weight (lbs)", min_value=MIN_WEIGHT_LBS, max_value=MAX_WEIGHT_LBS)
)
validate_systolic_bp = _as_public_validator(
    partial(_validate_int_range, field_label="Systolic BP", min_value=MIN_SYSTOLIC_BP, max_value=MAX_SYSTOLIC_BP)
)
validate_diastolic_bp = _as_public_validator(
    partial(_validate_int_range, field_label="Diastolic BP", min_value=MIN_DIASTOLIC_BP, max_value=MAX_DIASTOLIC_BP)
)


def validate_multi_value(items: list[str] | None, field_label: str) -> list[str] | None:
    """Used by PatientUpdate's multi-value field validators -- the payload
    arrives as list[str] from JSON directly, no comma-splitting needed (that
    only happens on the Excel ingestion path, via _split_multi_value_cell)."""
    if items is None:
        return None
    result, error = _validate_multi_value_items(items, field_label)
    if error:
        raise ValueError(error)
    return result


def _parse_date(raw: Any) -> date | None:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        text = raw.strip()
        for fmt in _DATE_STRING_FORMATS:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _years_before(base: date, years: int) -> date:
    try:
        return base.replace(year=base.year - years)
    except ValueError:
        # base is Feb 29 and base.year - years isn't a leap year.
        return base.replace(month=2, day=28, year=base.year - years)
